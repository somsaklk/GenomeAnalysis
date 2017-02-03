#!/usr/bin/env python
# 
# Author: somsak.lik@biotec.or.th
#
# Compare analysis outputs from two MAKER-derived genomes.
# Match homologous gene pairs which are aligned by blast results or whatever.
# (need to prepare gene match and homolog tables before run) 
# And save to Excel
#
# 2017/01/24

import os
from sys import stdout
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.cell import get_column_letter

# Specify needed file inputs
gene_table1 = 'OT_PacBIO8.out'
gene_table2 = 'DL1.out'
match_table = 'blastx_OT_PacBIO8_v_DL1.out_sum.out' # Summarized Blast results
												# Contains homolog gene names 
												# in column 1 and 2.
homolog_table = 'OT-DL1_homolog.tsv' # Homolog table compiled by Dr. Supatcha

f = open(gene_table1, 'r')
genes1 = f.readlines()
f.close()

g = open(gene_table2, 'r')
genes2 = g.readlines()
g.close()

m = open(match_table, 'r')
match_handle = m.readlines()
m.close()
dict1_2 = {}
dict2_1 = {}
for line in match_handle:
	line = line.rstrip()
	pair = line.split('\t')
	dict1_2[pair[0]] = pair[1]
	dict2_1[pair[1]] = pair[0]

h = open(homolog_table, 'r')
homolog_handle = h.readlines()
h.close()
h1 = homolog_handle[0].replace('\n', '').split('\t')[0]
h2 = homolog_handle[0].replace('\n', '').split('\t')[3]
del homolog_handle[0]
# Create homolog dict from the homolog_table file
homolog_dict = {}
for each in homolog_handle:
	each = each.replace('\n','\t')
	cols = each.split('\t')
	#set cols index according to main organism;
	#for example, set to cols[1] which is labelled 'OT' in file OT-DL1_homolog.tsv 
	homolog_dict[cols[2]] = [cols[0], cols[3]]

gene_table1 = gene_table1.replace('.out','')
gene_table2 = gene_table2.replace('.out', '')
xlsoutfile = 'sum_'+gene_table1+'-'+gene_table2 + '.xlsx'

# Excel formatting
wb = Workbook()
ws = wb.active
ws.title = "Genome_comparison"
boldft = Font(bold=True)
thin = Side(border_style="thin", color="555555")
double = Side(border_style="double", color="000000")
norm_border = Border(bottom=thin, left=thin, right=thin)
sep_border = Border(bottom=thin, left=double)
# For conditional formatting
notice_reverse = PatternFill("solid", fgColor="AD7FA8")
notice_forward = PatternFill("solid", fgColor="8AE234")
notice_absent = PatternFill("solid", fgColor="FCE94F")
notice_diff = PatternFill("solid", fgColor="DDEEDD")
warn_small = Font(color="4A4AA3")
warn_medium = Font(bold=True, color="BE6C00")
warn_big = Font(bold=True, color="EF4949")

############## Write header to excel
gene_header1 = genes1[0].rstrip()
gene_header2 = genes2[0].rstrip()
header_sec1 = gene_header1.split('\t') + ['Size_diff']
header_sec2 = gene_header2.split('\t')
header_sec3 = [h1,h2]
xlsheader = header_sec1 + header_sec2 + header_sec3 
sep1 = 1
sep2 = len(header_sec1)+1
sep3 = sep2 + len(header_sec2)
sep_col = [sep1,sep2,sep3]
for x in range(1, len(xlsheader)+1):
	d = ws.cell(row=1, column=x, value=xlsheader[x-1])
	d.font = Font(bold=True, color="FFFFFF")
	d.fill = PatternFill("solid", fgColor="666666")
	d.border = norm_border
	if x in sep_col:
		d.border = norm_border

del genes1[0]
del genes2[0]

shared = 0
unique1 = 0 #Number of gene found in Genome1, but not in Genome2
unique2 = 0 #Number of gene found in Genome2, but not in Genome1

row_startxls = 2 #start row for excel content

base_width = 3
col_width1 = []
col_width2 = []
for each1 in genes1:
	cols1 = each1.rstrip().split('\t')
	for i in range(3, len(cols1)):
		cols1[i] = int(cols1[i])
	gene_name = cols1[0]
	# Show running gene on screen
	stdout.write('\r%s' % gene_name)
	stdout.flush()
	col_width1.append(len(gene_name)+base_width)
	orientation1 = cols1[2]
	size1 = cols1[-1]
	row_toxls = []
	row_toxls.extend(cols1)
	if gene_name in dict1_2:
		shared+=1
		match_gene = dict1_2[gene_name]
		found = 0
		for each2 in genes2:
			cols2 = each2.rstrip().split('\t')
			col_width1.append(len(match_gene)+base_width)
			for i in range(3, len(cols2)):
				cols2[i] = int(cols2[i])
			if each2.startswith(match_gene):			
				orientation2 = cols2[2]
				size2 = cols2[-1]
				found+=1
				size_diff = size1-size2
				row_toxls.append(size_diff)
				if match_gene in homolog_dict.keys():
					row_toxls.extend(cols2)
					row_toxls.extend(homolog_dict[match_gene])
				else:
					row_toxls.extend(cols2)
					row_toxls += ['']*2
		if found == 0:
			size_diff = 'NA'
			row_toxls.append(size_diff)
			row_toxls += ['No match']*len(cols1)
			row_toxls += ['']*2
			orientation2 = 'No match'
			
	else:
		size_diff = 'NA'
		row_toxls.append(size_diff)
		row_toxls += ['No match']*len(cols1)
		row_toxls += ['']*2
		orientation2 = 'No match'
		unique1 +=1
	
	#Write to excel
	for x in range(1, len(row_toxls)+1):
		cell_result = ws.cell(row=row_startxls, column=x, value=row_toxls[x-1])
		cell_result.border = norm_border
		if x in sep_col:
			cell_result.border = sep_border
	# Emphasize gene name with bold font
	for gene_cell in (sep1, sep2):
		c = ws.cell(row=row_startxls, column=gene_cell)
		c.font = boldft
	# Label difference in gene orientation on chromosome strands
	if orientation1 != orientation2:
		for notice_cell in (sep1+2, sep2+2):
			d = ws.cell(row=row_startxls, column=notice_cell)
			if d.value == 'reverse':
				d.fill = notice_reverse
			elif d.value == 'forward':
				d.fill = notice_forward
			else:
				d.fill = notice_absent
	# Label size difference with font colors based on how big size_diff is.
	if isinstance(size_diff, (int,long)):
		for notice_cell in (sep1+5,sep1+6,sep2+5):
			e = ws.cell(row=row_startxls, column=notice_cell)
			if 0 < abs(size_diff) < 0.05*size1:
				e.font = warn_small
			elif 0.05*size1 <= abs(size_diff) < 0.25*size1:
				e.font = warn_medium
			elif abs(size_diff) >= 0.25*size1:
				e.font = warn_big
	diff_cell = ws.cell(row=row_startxls, column=sep1+len(cols1))	
	diff_cell.fill = notice_diff
		
	row_startxls+=1 #move excel cursor to next row
		
for each2 in genes2:
	cols2 = each2.rstrip().split('\t')
	for i in range(3, len(cols2)):
		cols2[i] = int(cols2[i])
	gene_name = cols2[0]
	# Show running gene on screen
	stdout.write('\r%s' % gene_name)
	stdout.flush()
	col_width2.append(len(gene_name)+base_width)
	row_toxls = []
	size_diff = 'NA'
	if not gene_name in dict2_1:
		orientation1 = 'No match'
		orientation2 = cols2[2]
		unique2+=1
		row_toxls += ['No match']*len(cols2)
		row_toxls.append(size_diff)
		if gene_name in homolog_dict.keys():
			row_toxls.extend(cols2)
			row_toxls.extend(homolog_dict[gene_name])
		else:
			row_toxls.extend(cols2)
			row_toxls += ['']*2
		#Write to excel
		for x in range(1, len(row_toxls)+1):
			cell_result = ws.cell(row=row_startxls, column=x, value=row_toxls[x-1])
			cell_result.border = norm_border
			if x in sep_col:
				cell_result.border = sep_border
		# Emphasize gene name with bold font
		for gene_cell in (sep1, sep2):
			c = ws.cell(row=row_startxls, column=gene_cell)
			c.font = boldft
		# Label difference in gene orientation on chromosome strands
		if orientation1 != orientation2:
			for notice_cell in (sep1+2, sep2+2):
				d = ws.cell(row=row_startxls, column=notice_cell)
				if d.value == 'reverse':
					d.fill = notice_reverse
				elif d.value == 'forward':
					d.fill = notice_forward
				else:
					d.fill = notice_absent
		# Label size difference with font colors based on how big size_diff is.
		if isinstance(size_diff, (int,long)):
			for notice_cell in (sep1+5,sep1+6,sep2+5):
				e = ws.cell(row=row_startxls, column=notice_cell)
				if 0 < abs(size_diff) < 0.05*size1:
					e.font = warn_small
				elif 0.05*size1 <= abs(size_diff) < 0.25*size1:
					e.font = warn_medium
				elif abs(size_diff) >= 0.25*size1:
					e.font = warn_big
		diff_cell = ws.cell(row=row_startxls, column=sep1+len(cols2))	
		diff_cell.fill = notice_diff
		
		row_startxls+=1 #move excel cursor to next row

# Widen columns with long word, e.g. gene name column
wide_cols = [sep1, sep2]
col_width = [max(col_width1), max(col_width2)]
for i, w in zip(wide_cols, col_width):
	ws.column_dimensions[get_column_letter(i)].width = w

########################################
# Insert new sheet, describing color code
ws2 = wb.create_sheet("Color_code")
ws2.column_dimensions['A'].width = 100
ws2['A1'] = "Gene located on forward strand in this genome, but on reverse or absent in the other"
ws2['A1'].fill = notice_forward
ws2['A2'] = "Gene located on reverse strand in this genome, but on forward or absent in the other"
ws2['A2'].fill = notice_reverse
ws2['A3'] = "Gene location absent in this genome, but present in the other"
ws2['A3'].fill = notice_absent

ws2['A4'] = "Size difference is between 0% and 5% of gene size"
ws2['A4'].font = warn_small
ws2['A5'] = "Size difference is between 5% and 25% of gene size"
ws2['A5'].font = warn_medium
ws2['A6'] = "Size difference is bigger than 25% of gene size"
ws2['A6'].font = warn_big
for legend_cell in ['A4','A5','A6']:
	ws2[legend_cell].fill = notice_diff
	
####################
# Insert new sheet, Summary by numbers
wssum = wb.create_sheet("By_numbers",0)
wssum.column_dimensions['A'].width = 55
wssum['A1'] = 'Total number of genes'
wssum['A1'].font = boldft
wssum['A2'] = 'Number of genes in %s' % gene_table1
wssum['B2'] = len(genes1)
wssum['A3'] = 'Number of genes in %s' % gene_table2
wssum['B3'] = len(genes2)

wssum['A5'] = 'BLAST: %s to %s' % (gene_table1, gene_table2)
wssum['A5'].font = boldft
wssum['A6'] = '%s genes that BLAST-HIT %s genes' % (gene_table1, gene_table2)
wssum['B6'] = shared
wssum['A7'] = '%s genes that DO NOT HIT %s genes' % (gene_table1, gene_table2)
wssum['B7'] = unique1
wssum['A8'] = '%s genes that ARE NOT HIT by %s genes' % (gene_table2, gene_table1)
wssum['B8'] = unique2

# Save from memory to Excel file
wb.save(xlsoutfile)

# Quick-look report on Terminal screen
os.system('cls' if os.name == 'nt' else 'clear')
print('''
Number of genes in %s = %d
Number of genes in %s = %d
''' % (gene_table1, len(genes1), gene_table2, len(genes2)))

print('''
Number of %s genes mapped to %s = %d
Number of genes unique to %s = %d
Number of genes unique to %s = %d
''' % (gene_table1, gene_table2, shared, gene_table1, unique1, gene_table2, unique2))
