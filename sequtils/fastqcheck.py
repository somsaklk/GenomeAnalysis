#!/usr/bin/env python 

from Bio import SeqIO
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

fastq_file = 'polished_assembly.fastq'
fastq_parser = SeqIO.parse(fastq_file, "fastq")

q = 3
q+=1

output_name = ".".join(fastq_file.split(".")[:-1])
#~ w = open(output_name+'_qualitycheck.tsv', 'w')

### Write header
header = ['contig','length(bp)']
for i in range(q):
	header.append('Q='+str(i))
header.append('percent_of_bases_with_Q<' + str(q-1))

#~ print("\t".join(header))
#~ w.write("\t".join(header) + '\n')

wb = Workbook()
ws = wb.active
ws.title = output_name

# Write header
boldft = Font(bold=True)
for x in range(1,len(header)+1):
	d = ws.cell(row=1, column=x, value=header[x-1])
	d.font = boldft
	
# Write legend
greyft = Font(color='55555555')
for y in range(q):
	legend = "'Q=%d': number of bases with phred scores equal %d (error prob. = %.2f)" % (y, y, 10.**(-y/10.)) 
	e = ws.cell(row=y+1, column=len(header)+2, value=legend)
	e.font = greyft
###
row_xlsx = 2
redft = Font(color=colors.RED)

for rec in fastq_parser:
	phred = rec.letter_annotations["phred_quality"]
	count_low_phred = []
	for i in range(q):
		n_base_phred = phred.count(i)
		count_low_phred.append(n_base_phred)
	
	seq_length = len(rec.seq)
	percent_low_phred = 100.*float(sum(count_low_phred))/seq_length
	
	row_to_write = [rec.id, str(seq_length)]
	for each in count_low_phred:
		row_to_write.append(str(each))
	
	row_to_write.append('%.2f%%' % percent_low_phred)
	#~ print("\t".join(row_to_write))
	#~ w.write("\t".join(row_to_write) + '\n')
	
	for x in range(1, len(row_to_write)+1):
		cell_result = ws.cell(row=row_xlsx, column=x, value=row_to_write[x-1])
		if percent_low_phred > 90.:
			cell_result.font = redft
	row_xlsx+=1

#~ w.close()
wb.save(output_name+'_qualitycheck.xlsx')
